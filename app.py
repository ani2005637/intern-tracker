from flask import Flask, request, jsonify, send_from_directory, session
import pymongo
from bson import ObjectId
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import os
import datetime
import openpyxl

app = Flask(__name__)

# Set secret key for sessions (secure default but customizable via environment)
app.secret_key = os.environ.get('SECRET_KEY', 'sanna_innovations_secret_key_129847')

# Enable CORS with credentials support so frontend can send/receive session cookies
CORS(app, supports_credentials=True)

# Connect to MongoDB
mongo_uri = os.environ.get('MONGO_URI')
if not mongo_uri:
    # Use default credentials provided by the user
    mongo_uri = 'mongodb+srv://srikakulaanirudh0506_db_user:e9gSM1PbSju6TYek@cluster0.3jbhs01.mongodb.net/?appName=Cluster0'

print("Connecting to MongoDB Atlas web service...")
client = pymongo.MongoClient(mongo_uri)
db = client['intern_tracker']

# Create indexes for high-performance data retrieval
print("Ensuring database indexes exist for fast loading...")
try:
    db.session_logs.create_index([("username", pymongo.ASCENDING), ("login_time", pymongo.DESCENDING)])
    db.tasks.create_index([("intern_name", pymongo.ASCENDING)])
    db.tasks.create_index([("assigned_by", pymongo.ASCENDING)])
    db.intern_logs.create_index([("intern_name", pymongo.ASCENDING), ("date_logged", pymongo.DESCENDING)])
    db.skills_log.create_index([("intern_name", pymongo.ASCENDING)])
    db.mentor_feedback.create_index([("intern_name", pymongo.ASCENDING)])
    db.notifications.create_index([("username", pymongo.ASCENDING), ("created_at", pymongo.DESCENDING)])
    db.announcements.create_index([("created_at", pymongo.DESCENDING)])
    db.direct_messages.create_index([("sender", pymongo.ASCENDING), ("recipient", pymongo.ASCENDING), ("created_at", pymongo.ASCENDING)])
    db.direct_messages.create_index([("recipient", pymongo.ASCENDING), ("created_at", pymongo.ASCENDING)])
    # Drop conflicting old index on attendance if it exists
    try:
        db.attendance.drop_index("username_1_date_1")
    except Exception:
        pass
    db.attendance.create_index([("username", pymongo.ASCENDING), ("year", pymongo.ASCENDING), ("month", pymongo.ASCENDING), ("day", pymongo.ASCENDING)], unique=True)
    print("Database indexes created/verified.")
except Exception as ie:
    print(f"Warning: Failed to ensure database indexes: {ie}")

# Startup migration to ensure all existing users have 'approved: True' and 'restricted: False'
try:
    print("Running startup migration to approve existing users and initialize restriction state...")
    db.users.update_many({"approved": {"$exists": False}}, {"$set": {"approved": True}})
    db.users.update_many({"restricted": {"$exists": False}}, {"$set": {"restricted": False}})
    print("Startup migration completed successfully.")
except Exception as me:
    print(f"Warning: Startup migration failed: {me}")

def sync_attendance_on_boot():
    try:
        import openpyxl
        import re

        NAME_TO_USERNAME = {
            "srinivas": "empsi01",
            "jyothi": "empsi02",
            "sowmya": "empsi03",
            "naveen kumar": "empsi04",
            "nikki": "empsi05",
            "mavuri anand kumar": "empsi06",
            "riya biswas dasgupta": "empsi07",
            "adarsh": "empsi08",
            "jakka balaji mahendra": "empsi09",
            "bhuvan sai adithya": "empsi10",
            "m.shalini sree": "empsi11",
            "vansh goyal": "empsi12",
            "srikakula anirudh": "empsi13"
        }

        # Check if attendance database already has records
        if db.attendance.count_documents({}) > 0:
            print("Attendance collection already populated. Skipping boot sync to preserve live edits.")
            return

        print("Starting boot sync for multiple months...")
        
        # Clear existing attendance records to ensure clean sync on first boot
        db.attendance.delete_many({})
        print("Cleared existing attendance database records.")

        sync_jobs = [
            ("jan-and-feb.xlsx", "Jan", 2026, 1),
            ("jan-and-feb.xlsx", "Feb", 2026, 2),
            ("march-april-may.xlsx", "Mar", 2026, 3),
            ("march-april-may.xlsx", "Apr", 2026, 4),
            ("march-april-may.xlsx", "May 2026", 2026, 5),
            ("attendance.xlsx", None, 2026, 6) # None defaults to active sheet
        ]

        count_users = 0
        count_attendance = 0

        for filename, sheetname, year, month in sync_jobs:
            excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
            if not os.path.exists(excel_path):
                print(f"File {filename} not found. Skipping.")
                continue

            print(f"Syncing {filename} -> {sheetname or 'Active Sheet'} for {year}-{month}...")
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            sheet = wb[sheetname] if sheetname else wb.active

            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                print(f"Sheet {sheetname or 'Active'} in {filename} is empty or lacks rows.")
                continue

            # Build header mapping
            headers = rows[0]
            header_map = {}
            day_columns = {} # day_number -> column_index

            for idx, h in enumerate(headers):
                if h is None:
                    continue
                if isinstance(h, (datetime.datetime, datetime.date)):
                    day_columns[h.day] = idx
                elif isinstance(h, str):
                    h_clean = h.strip().lower()
                    header_map[h_clean] = idx
                    
                    # Regex to match "day 1", "day_1", "day1"
                    m = re.match(r'^(?:day)\s*(\d+)$', h_clean)
                    if m:
                        day_columns[int(m.group(1))] = idx
                elif isinstance(h, int):
                    day_columns[h] = idx

            # Safely get header index
            def get_header_idx(keys):
                for k in keys:
                    if k in header_map:
                        return header_map[k]
                return None

            emp_no_idx = get_header_idx(['emp no', 'emp_no', 'employee no', 'emp id', 'emp_id'])
            name_idx = get_header_idx(['employee name', 'employee_name', 'name'])
            designation_idx = get_header_idx(['designation', 'title'])
            salary_idx = get_header_idx(['salary', 'monthly salary', 'gross salary', 'gross_salary'])
            doj_idx = get_header_idx(['date of joining', 'date_of_joining', 'doj'])
            active_idx = get_header_idx(['active', 'active status', 'status'])

            if emp_no_idx is None or name_idx is None:
                print(f"Critical headers 'Emp No' or 'Employee Name' not found in {filename} -> {sheetname or 'Active'}.")
                continue

            for r_idx in range(1, len(rows)):
                row = rows[r_idx]
                if not row or emp_no_idx >= len(row) or row[emp_no_idx] is None:
                    continue

                emp_name = str(row[name_idx]).strip()
                norm_name = " ".join(emp_name.lower().split())
                canonical_uname = NAME_TO_USERNAME.get(norm_name)
                
                emp_no = str(row[emp_no_idx]).strip().lower()
                if canonical_uname:
                    emp_no = canonical_uname

                if not emp_no or emp_no in ['total', 'totals']:
                    continue
                designation = str(row[designation_idx]).strip() if (designation_idx is not None and designation_idx < len(row) and row[designation_idx] is not None) else "Employee"
                
                salary = 0.0
                if salary_idx is not None and salary_idx < len(row) and row[salary_idx] is not None:
                    try:
                        salary = float(row[salary_idx])
                    except ValueError:
                        pass

                doj = ""
                if doj_idx is not None and doj_idx < len(row) and row[doj_idx] is not None:
                    val = row[doj_idx]
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        doj = val.strftime("%Y-%m-%d")
                    else:
                        val_str = str(val).strip()
                        m = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$', val_str)
                        if m:
                            doj = f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
                        else:
                            doj = val_str

                active = True
                if active_idx is not None and active_idx < len(row) and row[active_idx] is not None:
                    val_str = str(row[active_idx]).strip().lower()
                    if val_str in ['no', 'false', '0', 'inactive', 'restricted']:
                        active = False

                update_fields = {
                    "full_name": emp_name,
                    "title": designation,
                    "salary": salary,
                    "date_of_joining": doj,
                    "active": active,
                    "approved": active,
                    "restricted": not active,
                    "is_employee_profile": True
                }

                user_doc = db.users.find_one({"username": emp_no})
                if not user_doc:
                    db.users.insert_one({
                        "username": emp_no,
                        "password_hash": generate_password_hash(emp_no),
                        "role": "Employee",
                        "email": f"{emp_no}@sannainnovations.com",
                        "created_at": datetime.datetime.utcnow(),
                        **update_fields
                    })
                else:
                    db.users.update_one({"username": emp_no}, {"$set": update_fields})

                count_users += 1

                # Store month-specific designation and salary
                db.monthly_payroll_profiles.update_one(
                    {"username": emp_no, "year": year, "month": month},
                    {"$set": {
                        "title": designation,
                        "salary": salary
                    }},
                    upsert=True
                )

                for d, col_idx in day_columns.items():
                    if col_idx >= len(row):
                        continue
                    status_val = row[col_idx]
                    if status_val is not None:
                        status_val = str(status_val).strip().upper()
                        if status_val in ['P', 'A', 'HD', 'CL', 'SL', 'LOP', 'H']:
                            db.attendance.update_one(
                                {"username": emp_no, "year": year, "month": month, "day": d},
                                {"$set": {"status": status_val}},
                                upsert=True
                            )
                            count_attendance += 1

        print(f"Sync complete. Synchronized {count_users} employee rows and {count_attendance} attendance logs.")

    except Exception as e:
        print(f"Error during boot attendance sync: {e}")

sync_attendance_on_boot()

# Helper to serialize MongoDB documents (converting ObjectId to string 'id')
def serialize(doc):
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize(d) for d in doc]
    
    doc = dict(doc)
    if '_id' in doc:
        doc['id'] = str(doc['_id'])
        del doc['_id']
    return doc

def validate_log_date(date_str):
    if not date_str:
        return "Log date is required"
    try:
        log_dt = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        today_dt = datetime.date.today()
        yesterday_dt = today_dt - datetime.timedelta(days=1)
        if log_dt < yesterday_dt or log_dt > today_dt:
            return "Daily log date must be yesterday or today."
    except ValueError:
        return "Invalid log date format. Expected YYYY-MM-DD"
    return None

def validate_future_or_today_date(date_str, field_name):
    if not date_str:
        return None
    try:
        dt = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        today_dt = datetime.date.today()
        if dt < today_dt:
            return f"{field_name} cannot be in the past"
    except ValueError:
        return f"Invalid {field_name} format. Expected YYYY-MM-DD"
    return None

# ----------------- AUTHENTICATION MIDDLEWARE HELPERS -----------------
def get_logged_in_user():
    user_id = session.get('user_id')
    if not user_id:
        return None
    try:
        user = db.users.find_one({"_id": ObjectId(user_id)})
        return user
    except Exception:
        return None

# Helper to find usernames of all subordinate users (strict hierarchical access)
def get_subordinate_usernames(user):
    role_weights = {"Intern": 1, "Employee": 2, "Manager": 3, "Admin": 4}
    user_weight = role_weights.get(user['role'], 0)
    
    if user['role'] in ['Admin', 'Guest']:
        return None
        
    # Find all roles with a weight strictly lower than user's weight
    sub_roles = [role for role, weight in role_weights.items() if weight < user_weight]
    
    # Query usernames of those roles
    sub_users = list(db.users.find({"role": {"$in": sub_roles}}, {"username": 1}))
    sub_usernames = [u['username'] for u in sub_users]
    
    # Always include the user's own username so they can see their own data
    sub_usernames.append(user['username'])
    return sub_usernames


@app.before_request
def restrict_guest_writes():
    if request.method in ['POST', 'PUT', 'DELETE']:
        if request.path in ['/login', '/logout', '/signup']:
            return None
        user = get_logged_in_user()
        if user and user.get('role') == 'Guest':
            return jsonify({"error": "Forbidden: Guest account is read-only"}), 403
    return None


# ----------------- FRONTEND ROUTE -----------------
@app.route('/')
def serve_frontend():
    response = send_from_directory('.', 'index.html')
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# ----------------- AUTHENTICATION API ENDPOINTS -----------------
@app.route('/signup', methods=['POST'])
def signup():
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    username = data.get('username', '').strip().lower()
    password = data.get('password', '')
    full_name = data.get('full_name', '').strip()
    role = data.get('role', 'Intern')  # 'Admin', 'Manager', 'Employee', 'Intern'
    email = data.get('email', '').strip()
    title = data.get('title', '').strip()
    employee_id = data.get('employee_id', '').strip()

    if not username or not password or not full_name or not role or not email or not title:
        return jsonify({"error": "All fields are required (except Employee ID)"}), 400

    if role not in ['Admin', 'Manager', 'Employee', 'Intern']:
        return jsonify({"error": "Invalid role specified"}), 400

    # Check if username exists
    existing = db.users.find_one({"username": username})
    if existing:
        return jsonify({"error": "Username already taken"}), 400

    hashed_pw = generate_password_hash(password)
    try:
        db.users.insert_one({
            "username": username,
            "password_hash": hashed_pw,
            "full_name": full_name,
            "role": role,
            "email": email,
            "title": title,
            "employee_id": employee_id,
            "approved": False,
            "created_at": datetime.datetime.utcnow()
        })
        
        # Notify all Admin users about the new pending registration
        try:
            admins = list(db.users.find({"role": "Admin"}))
            for admin_user in admins:
                db.notifications.insert_one({
                    "username": admin_user['username'],
                    "message": f"New registration pending approval: {full_name} ({role})",
                    "type": "registration_pending",
                    "created_at": datetime.datetime.utcnow(),
                    "read": False
                })
        except Exception as ne:
            print(f"Failed to create admin notification on signup: {ne}")
            
    except Exception as e:
        return jsonify({"error": f"Signup failed: {str(e)}"}), 500

    return jsonify({"message": "Registration successful! Your account is pending administrator approval."}), 201

@app.route('/users/create', methods=['POST'])
def create_user_by_supervisor():
    user = get_logged_in_user()
    if not user or user['role'] not in ['Admin', 'Manager']:
        return jsonify({"error": "Unauthorized: Access denied"}), 401
        
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    username = data.get('username', '').strip().lower()
    password = data.get('password', '')
    full_name = data.get('full_name', '').strip()
    role = data.get('role', '').strip()
    email = data.get('email', '').strip()
    title = data.get('title', '').strip()
    employee_id = data.get('employee_id', '').strip()

    if not username or not password or not full_name or not role or not email or not title:
        return jsonify({"error": "All fields are required (except Employee ID)"}), 400

    if role not in ['Admin', 'Manager', 'Employee', 'Intern', 'Guest']:
        return jsonify({"error": "Invalid role specified"}), 400

    # Hierarchical role check
    if user['role'] == 'Manager' and role not in ['Employee', 'Intern']:
        return jsonify({"error": "Forbidden: Managers can only create Employee or Intern accounts"}), 403

    # Check if username exists
    existing = db.users.find_one({"username": username})
    if existing:
        return jsonify({"error": "Username already taken"}), 400

    hashed_pw = generate_password_hash(password)
    try:
        db.users.insert_one({
            "username": username,
            "password_hash": hashed_pw,
            "full_name": full_name,
            "role": role,
            "email": email,
            "title": title,
            "employee_id": employee_id,
            "approved": True, # Pre-approved since created by supervisor
            "restricted": False,
            "created_at": datetime.datetime.utcnow()
        })
        return jsonify({"message": f"Successfully created user account for {full_name} ({role})!"}), 201
    except Exception as e:
        return jsonify({"error": f"Failed to create user: {str(e)}"}), 500


@app.route('/login', methods=['POST'])
def login():
    data = request.json
    if not data:
        return jsonify({"error": "No credentials provided"}), 400

    username = data.get('username', '').strip().lower()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    user = db.users.find_one({
        "$or": [
            {"username": username},
            {"alias_username": username}
        ]
    })
    if not user:
        return jsonify({"error": "Invalid credentials"}), 401

    if not check_password_hash(user['password_hash'], password):
        return jsonify({"error": "Invalid credentials"}), 401

    # Check if user account is approved
    if not user.get('approved', False):
        return jsonify({"error": "Your account is pending administrator approval. Please contact an admin."}), 403

    # Check if user account is restricted
    if user.get('restricted', False):
        return jsonify({"error": "Your account has been temporarily restricted by an administrator or manager. Please contact support."}), 403

    # Store user identity details inside the session
    session['user_id'] = str(user['_id'])
    session['username'] = user['username']
    session['role'] = user['role']
    session['full_name'] = user['full_name']

    # Close any orphaned active sessions for this user
    try:
        db.session_logs.update_many(
            {"username": user['username'], "logout_time": None},
            {"$set": {"logout_time": datetime.datetime.utcnow()}}
        )
    except Exception as e:
        print(f"Failed to close orphaned sessions: {e}")

    # Record login activity log
    try:
        db.session_logs.insert_one({
            "username": user['username'],
            "full_name": user['full_name'],
            "role": user['role'],
            "login_time": datetime.datetime.utcnow(),
            "logout_time": None
        })
    except Exception as e:
        print(f"Failed to record login session: {e}")

    return jsonify({
        "message": "Login successful!",
        "user": {
            "username": user['username'],
            "full_name": user['full_name'],
            "role": user['role'],
            "employee_id": user.get('employee_id', '')
        }
    })


@app.route('/logout', methods=['POST'])
def logout():
    user = get_logged_in_user()
    if user:
        try:
            # Close all active open sessions for this user
            res = db.session_logs.update_many(
                {"username": user['username'], "logout_time": None},
                {"$set": {"logout_time": datetime.datetime.utcnow()}}
            )
            # If no sessions were modified, fallback to updating the latest session's logout time
            if res.modified_count == 0:
                latest_any = db.session_logs.find_one(
                    {"username": user['username']},
                    sort=[("login_time", pymongo.DESCENDING)]
                )
                if latest_any:
                    db.session_logs.update_one(
                        {"_id": latest_any["_id"]},
                        {"$set": {"logout_time": datetime.datetime.utcnow()}}
                    )
        except Exception as e:
            print(f"Failed to record logout session: {e}")

    session.clear()
    return jsonify({"message": "Logged out successfully!"})


@app.route('/current_user', methods=['GET'])
def current_user():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify({
        "username": user['username'],
        "full_name": user['full_name'],
        "role": user['role'],
        "email": user.get('email', ''),
        "employee_id": user.get('employee_id', ''),
        "title": user.get('title', '')
    })



@app.route('/users/<username>/profile', methods=['PUT'])
def update_user_profile(username):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if user['username'].lower() != username.lower() and user['role'] != 'Admin':
        return jsonify({"error": "Forbidden: You can only edit your own profile"}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    update_fields = {}
    if 'full_name' in data and data['full_name'].strip():
        update_fields['full_name'] = data['full_name'].strip()
    if 'email' in data and data['email'].strip():
        update_fields['email'] = data['email'].strip()
    if 'employee_id' in data:
        update_fields['employee_id'] = data['employee_id'].strip()
    if 'password' in data and data['password']:
        update_fields['password_hash'] = generate_password_hash(data['password'])

    if not update_fields:
        return jsonify({"error": "No fields to update"}), 400

    try:
        db.users.update_one({"username": username}, {"$set": update_fields})
        if 'full_name' in update_fields:
            new_name = update_fields['full_name']
            db.logs.update_many({"intern_name": username}, {"$set": {"intern_fullname": new_name}})
            db.tasks.update_many({"intern_name": username}, {"$set": {"intern_fullname": new_name}})
            db.leave_requests.update_many({"username": username}, {"$set": {"full_name": new_name}})
            db.session_logs.update_many({"username": username}, {"$set": {"full_name": new_name}})

        return jsonify({"message": "Profile updated successfully"})
    except Exception as e:
        return jsonify({"error": f"Failed to update profile: {str(e)}"}), 500


# Fetch users list dynamically (useful for task assignment dropdowns)
@app.route('/users', methods=['GET'])
def get_users():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    role_filter = request.args.get('role')
    query = {}
    if role_filter:
        query['role'] = role_filter

    # Fetch users
    users = list(db.users.find(query).sort("full_name", 1))
    
    # Filter out test users
    users = [u for u in users if not (
        'test' in u['username'].lower() or 
        'test' in u.get('full_name', '').lower() or 
        'pending' in u.get('full_name', '').lower() or
        u['username'].startswith('mgr_') or 
        u['username'].startswith('emp_') or 
        u['username'].startswith('int_')
    )]
    
    # Optimize N+1 query: Fetch the latest session for all users in a single aggregation query
    try:
        pipeline = [
            {"$sort": {"login_time": -1}},
            {"$group": {
                "_id": "$username",
                "login_time": {"$first": "$login_time"},
                "logout_time": {"$first": "$logout_time"}
            }}
        ]
        sessions = {s["_id"]: s for s in db.session_logs.aggregate(pipeline)}
    except Exception as ae:
        print(f"Warning: Session aggregation failed: {ae}")
        sessions = {}

    for u in users:
        latest_session = sessions.get(u['username'])
        if latest_session:
            u['last_login'] = latest_session.get('login_time')
            u['last_logout'] = latest_session.get('logout_time')
            if latest_session.get("logout_time") is None:
                u['status'] = "Available"
            else:
                u['status'] = "Logged Out"
        else:
            u['last_login'] = None
            u['last_logout'] = None
            u['status'] = "Logged Out"

    return jsonify(serialize(users))


@app.route('/users/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if user['role'] not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Only Admins and Managers can remove users"}), 403

    try:
        target_user = db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        return jsonify({"error": "Invalid user ID format"}), 400

    if not target_user:
        return jsonify({"error": "User not found"}), 404

    # Prevent self-deletion
    if str(target_user['_id']) == str(user['_id']):
        return jsonify({"error": "You cannot remove yourself from the portal"}), 400

    role_weights = {
        "Intern": 1,
        "Employee": 2,
        "Manager": 3,
        "Admin": 4
    }

    user_weight = role_weights.get(user['role'], 0)
    target_weight = role_weights.get(target_user['role'], 0)

    # Enforce hierarchical access control: Rank(A) > Rank(B) (Bypassed for Admin)
    if user['role'] != 'Admin' and user_weight <= target_weight:
        return jsonify({"error": "Forbidden: You can only remove users with a lower role rank than yours"}), 403

    try:
        # Delete user
        db.users.delete_one({"_id": ObjectId(user_id)})
        
        # Cascade delete target user's records
        db.intern_logs.delete_many({"intern_name": target_user['username']})
        db.tasks.delete_many({"intern_name": target_user['username']})
        db.skills_log.delete_many({"intern_name": target_user['username']})
        db.mentor_feedback.delete_many({"intern_name": target_user['username']})
        
    except Exception as e:
        return jsonify({"error": f"Failed to remove user: {str(e)}"}), 500

    return jsonify({"message": f"User {target_user['full_name']} removed successfully!"})


# ----------------- EMPLOYEE DIRECTORY & ATTENDANCE MATRIX ENDPOINTS -----------------

@app.route('/api/employees', methods=['GET'])
def get_employees():
    user = get_logged_in_user()
    if not user or user.get('role') not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Access restricted to Admin and Manager"}), 403

    now = datetime.datetime.utcnow()
    year = request.args.get('year', default=now.year, type=int)
    month = request.args.get('month', default=now.month, type=int)

    # Fetch users who are employee profiles (seeded from spreadsheet or manually added)
    users_list = list(db.users.find({"is_employee_profile": True}))

    # Sort alphabetically by username (places empsi01 first)
    users_list.sort(key=lambda x: x['username'])

    for u in users_list:
        uname = u['username']
        
        # CL rolls over month-to-month. Earns 1 CL per month.
        cl_allocated = float(month) * 1.0
        cl_used = db.attendance.count_documents({
            "username": uname,
            "year": year,
            "month": {"$lte": month},
            "status": "CL"
        })
        cl_remaining = max(0.0, cl_allocated - float(cl_used))

        # SL is 1 day per month and does NOT get forwarded.
        sl_allocated = 1.0
        sl_used = db.attendance.count_documents({
            "username": uname,
            "year": year,
            "month": month,
            "status": "SL"
        })
        sl_remaining = max(0.0, sl_allocated - float(sl_used))

        u['cl_remaining'] = cl_remaining
        u['sl_remaining'] = sl_remaining

    return jsonify(serialize(users_list))


@app.route('/api/employees/<username>', methods=['PUT'])
def update_employee(username):
    user = get_logged_in_user()
    if not user or user.get('role') not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Access restricted to Admin and Manager"}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    target_user = db.users.find_one({"username": username.lower().strip()})
    if not target_user:
        return jsonify({"error": "User not found"}), 404

    # Prevent deactivating other Managers or Admins if active status/details
    if user['role'] == 'Manager' and target_user['role'] in ['Admin', 'Manager'] and target_user['username'] != user['username']:
        return jsonify({"error": "Forbidden: Managers cannot update other Managers or Admins"}), 403

    designation = data.get('designation')
    salary = data.get('salary')
    date_of_joining = data.get('date_of_joining')
    active = data.get('active')

    update_fields = {}
    if designation is not None:
        update_fields["title"] = str(designation).strip()
    if salary is not None:
        try:
            update_fields["salary"] = float(salary)
        except ValueError:
            return jsonify({"error": "Invalid salary format"}), 400
    if date_of_joining is not None:
        update_fields["date_of_joining"] = str(date_of_joining).strip()
    if active is not None:
        is_active = bool(active)
        update_fields["active"] = is_active
        update_fields["approved"] = is_active
        update_fields["restricted"] = not is_active

    if update_fields:
        db.users.update_one({"username": username.lower().strip()}, {"$set": update_fields})

    return jsonify({"message": f"Employee {username} updated successfully!"})


@app.route('/api/employees', methods=['POST'])
def add_employee():
    user = get_logged_in_user()
    if not user or user.get('role') not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Access restricted to Admin and Manager"}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    username = str(data.get('username', '')).strip().lower()
    fullname = str(data.get('fullname', '')).strip()
    designation = str(data.get('designation', '')).strip()
    salary_val = data.get('salary')
    date_of_joining = str(data.get('date_of_joining', '')).strip()
    active = bool(data.get('active', True))

    if not username or not fullname or not designation or salary_val is None or not date_of_joining:
        return jsonify({"error": "Missing required fields"}), 400

    try:
        salary = float(salary_val)
    except ValueError:
        return jsonify({"error": "Invalid salary format"}), 400

    # Check if user already exists
    existing_user = db.users.find_one({"username": username})
    if existing_user:
        return jsonify({"error": "Employee Code / Username already exists"}), 400

    # Insert employee
    db.users.insert_one({
        "username": username,
        "password_hash": generate_password_hash(username), # default password is username
        "role": "Employee",
        "email": f"{username}@sannainnovations.com",
        "created_at": datetime.datetime.utcnow(),
        "full_name": fullname,
        "title": designation,
        "salary": salary,
        "date_of_joining": date_of_joining,
        "active": active,
        "approved": active,
        "restricted": not active,
        "is_employee_profile": True
    })

    # Ensure leave balances
    ensure_leave_balances(username)

    return jsonify({"message": f"Employee {fullname} added successfully!"})


@app.route('/api/attendance/overview', methods=['GET'])
def get_attendance_overview():
    user = get_logged_in_user()
    if not user or user.get('role') not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Access restricted to Admin and Manager"}), 403

    year_str = request.args.get('year')
    month_str = request.args.get('month')
    if not year_str or not month_str:
        return jsonify({"error": "Missing year or month parameters"}), 400

    try:
        year = int(year_str)
        month = int(month_str)
    except ValueError:
        return jsonify({"error": "Invalid year or month format"}), 400

    import calendar
    # Calculate days in month
    _, num_days = calendar.monthrange(year, month)

    # Calculate default working days (excluding Sundays)
    default_working_days = 0
    for d in range(1, num_days + 1):
        dt = datetime.date(year, month, d)
        if dt.weekday() != 6:  # 6 is Sunday
            default_working_days += 1

    working_days_override_str = request.args.get('total_working_days')
    total_working_days = default_working_days
    if working_days_override_str:
        try:
            total_working_days = int(working_days_override_str)
            if total_working_days <= 0:
                total_working_days = default_working_days
        except ValueError:
            pass

    # Fetch employees list (same sorting and filtering as GET /api/employees)
    employees = list(db.users.find({"is_employee_profile": True}))
    employees.sort(key=lambda x: x['username'])

    # Fetch monthly payroll profiles
    profiles_cursor = db.monthly_payroll_profiles.find({"year": year, "month": month})
    monthly_profiles = {p['username']: p for p in profiles_cursor}

    for emp in employees:
        uname = emp['username']
        m_profile = monthly_profiles.get(uname, {})
        if 'title' in m_profile:
            emp['title'] = m_profile['title']
        if 'salary' in m_profile:
            emp['salary'] = m_profile['salary']

    # Fetch attendance logs
    logs_cursor = db.attendance.find({"year": year, "month": month})
    attendance_map = {} # username -> day -> status
    for log in logs_cursor:
        uname = log['username']
        day = log['day']
        status = log['status']
        if uname not in attendance_map:
            attendance_map[uname] = {}
        attendance_map[uname][day] = status

    summary = {}
    for emp in employees:
        uname = emp['username']
        emp_logs = attendance_map.get(uname, {})
        
        # Count statuses
        counts = {"P": 0, "A": 0, "HD": 0, "CL": 0, "SL": 0, "LOP": 0, "H": 0}
        for d in range(1, num_days + 1):
            status = emp_logs.get(d)
            if status in counts:
                counts[status] += 1

        monthly_salary = emp.get('salary', 0.0)
        daily_rate = 0.0
        deductions = 0.0
        net_payable = monthly_salary

        if monthly_salary > 0:
            daily_rate = monthly_salary / total_working_days
            lop_days = counts["LOP"] + counts["A"]
            deductions = daily_rate * lop_days
            net_payable = max(0.0, monthly_salary - deductions)

        summary[uname] = {
            "P": counts["P"],
            "A": counts["A"],
            "HD": counts["HD"],
            "CL": counts["CL"],
            "SL": counts["SL"],
            "H": counts["H"],
            "LOP": counts["LOP"],
            "daily_rate": round(daily_rate, 2),
            "deductions": round(deductions, 2),
            "net_payable": round(net_payable, 2)
        }

    return jsonify({
        "employees": serialize(employees),
        "attendance_matrix": attendance_map,
        "summary": summary,
        "days_in_month": num_days,
        "total_working_days": total_working_days
    })


@app.route('/api/attendance', methods=['POST'])
def update_attendance_cell():
    user = get_logged_in_user()
    if not user or user.get('role') not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Access restricted to Admin and Manager"}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    username = data.get('username')
    year = data.get('year')
    month = data.get('month')
    day = data.get('day')
    status = data.get('status')

    if not username or year is None or month is None or day is None:
        return jsonify({"error": "Missing required fields"}), 400

    try:
        year = int(year)
        month = int(month)
        day = int(day)
    except ValueError:
        return jsonify({"error": "Invalid numeric values"}), 400

    status = str(status).strip().upper()
    if status == "":
        db.attendance.delete_one({"username": username.lower().strip(), "year": year, "month": month, "day": day})
        return jsonify({"message": "Attendance cell cleared successfully!"})

    if status not in ['P', 'A', 'HD', 'CL', 'SL', 'LOP', 'H']:
        return jsonify({"error": "Invalid status value"}), 400

    target_user = db.users.find_one({"username": username.lower().strip()})
    if not target_user:
        return jsonify({"error": "Target user not found"}), 404

    db.attendance.update_one(
        {"username": username.lower().strip(), "year": year, "month": month, "day": day},
        {"$set": {"status": status}},
        upsert=True
    )

    return jsonify({"message": "Attendance cell updated successfully!"})


@app.route('/api/attendance/export', methods=['GET'])
def export_attendance_to_excel():
    user = get_logged_in_user()
    if not user or user.get('role') not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Access restricted to Admin and Manager"}), 403

    year_str = request.args.get('year')
    month_str = request.args.get('month')
    if not year_str or not month_str:
        return jsonify({"error": "Missing year or month parameters"}), 400

    try:
        year = int(year_str)
        month = int(month_str)
    except ValueError:
        return jsonify({"error": "Invalid year or month format"}), 400

    import calendar
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    _, num_days = calendar.monthrange(year, month)

    # Calculate default working days (excluding Sundays)
    default_working_days = 0
    for d in range(1, num_days + 1):
        dt = datetime.date(year, month, d)
        if dt.weekday() != 6:  # 6 is Sunday
            default_working_days += 1

    working_days_override_str = request.args.get('total_working_days')
    total_working_days = default_working_days
    if working_days_override_str:
        try:
            total_working_days = int(working_days_override_str)
            if total_working_days <= 0:
                total_working_days = default_working_days
        except ValueError:
            pass

    # Fetch employees list
    employees = list(db.users.find({"is_employee_profile": True}))
    employees.sort(key=lambda x: x['username'])

    # Fetch monthly payroll profiles
    profiles_cursor = db.monthly_payroll_profiles.find({"year": year, "month": month})
    monthly_profiles = {p['username']: p for p in profiles_cursor}

    for emp in employees:
        uname = emp['username']
        m_profile = monthly_profiles.get(uname, {})
        if 'title' in m_profile:
            emp['title'] = m_profile['title']
        if 'salary' in m_profile:
            emp['salary'] = m_profile['salary']

    # Fetch attendance logs
    logs_cursor = db.attendance.find({"year": year, "month": month})
    attendance_map = {} # username -> day -> status
    for log in logs_cursor:
        uname = log['username']
        day = log['day']
        status = log['status']
        if uname not in attendance_map:
            attendance_map[uname] = {}
        attendance_map[uname][day] = status

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance_{calendar.month_abbr[month]}_{year}"

    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True

    # Colors
    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid") # Dark indigo
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Light grey
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Fonts
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    title_font = Font(name="Arial", size=14, bold=True, color="1E1B4B")
    subtitle_font = Font(name="Arial", size=10, italic=True)

    # Borders
    thin_border = Border(
        left=Side(style='thin', color="E4E4E7"),
        right=Side(style='thin', color="E4E4E7"),
        top=Side(style='thin', color="E4E4E7"),
        bottom=Side(style='thin', color="E4E4E7")
    )

    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Title Block
    month_name = calendar.month_name[month]
    ws.cell(row=1, column=1, value=f"Sanna Innovations - Attendance & Payroll Matrix").font = title_font
    ws.cell(row=2, column=1, value=f"Period: {month_name} {year} | Total Working Days: {total_working_days}").font = subtitle_font
    
    # Row 4: Column Headers
    headers = [
        "Employee Code", "Employee Name", "Designation", "Monthly Salary"
    ]
    for d in range(1, num_days + 1):
        headers.append(f"Day {d}")
        
    headers.extend([
        "Presents (P)", "Absents (A)", "Half Days (HD)", "Casual Leaves (CL)", "Sick Leaves (SL)", "Holidays (H)", "LOP Days", "LOP Deduction", "Net Payable"
    ])

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Populate Data
    start_row = 5
    for idx, emp in enumerate(employees):
        row_num = start_row + idx
        uname = emp['username']
        emp_logs = attendance_map.get(uname, {})
        
        # Zebra striping
        row_fill = zebra_fill if idx % 2 == 1 else white_fill

        ws.cell(row=row_num, column=1, value=uname.upper()).alignment = center_align
        ws.cell(row=row_num, column=2, value=emp['full_name']).alignment = left_align
        ws.cell(row=row_num, column=3, value=emp.get('title', 'Employee')).alignment = left_align
        
        # Monthly Salary
        sal_cell = ws.cell(row=row_num, column=4, value=emp.get('salary', 0.0))
        sal_cell.alignment = right_align
        sal_cell.number_format = '[$₹-4009] #,##0.00'

        # Set borders & fonts for first 4 cells
        for c in range(1, 5):
            ws.cell(row=row_num, column=c).font = data_font
            ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=c).fill = row_fill

        # Day columns (Col 5 to Col 5 + num_days - 1)
        counts = {"P": 0, "A": 0, "HD": 0, "CL": 0, "SL": 0, "LOP": 0, "H": 0}
        for d in range(1, num_days + 1):
            status = emp_logs.get(d)
            col_num = 4 + d
            cell = ws.cell(row=row_num, column=col_num, value=status if status else None)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = row_fill
            
            if status in counts:
                counts[status] += 1

        # Summary columns
        sum_cols_start = 4 + num_days + 1
        summary_values = [
            counts["P"], counts["A"], counts["HD"], counts["CL"], counts["SL"], counts["H"],
            counts["LOP"] + counts["A"]
        ]
        
        for c_offset, val in enumerate(summary_values):
            col_num = sum_cols_start + c_offset
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = row_fill

        # LOP Deduction & Net Payable Formulas
        salary_cell_coord = f"D{row_num}"
        lop_days_cell_coord = f"{get_column_letter(sum_cols_start + 6)}{row_num}" # LOP Days column
        
        lop_ded_col_letter = get_column_letter(sum_cols_start + 7)
        
        # LOP Deduction formula: = (Salary / working_days) * LOP_days
        lop_ded_formula = f"=IF({salary_cell_coord}>0, ({salary_cell_coord}/{total_working_days})*{lop_days_cell_coord}, 0)"
        lop_ded_cell = ws.cell(row=row_num, column=sum_cols_start + 7, value=lop_ded_formula)
        lop_ded_cell.alignment = right_align
        lop_ded_cell.number_format = '[$₹-4009] #,##0.00'
        lop_ded_cell.font = data_font
        lop_ded_cell.border = thin_border
        lop_ded_cell.fill = row_fill

        # Net Payable formula: = MAX(0, Salary - LOP Deduction)
        net_pay_formula = f"=MAX(0, {salary_cell_coord}-{lop_ded_col_letter}{row_num})"
        net_pay_cell = ws.cell(row=row_num, column=sum_cols_start + 8, value=net_pay_formula)
        net_pay_cell.alignment = right_align
        net_pay_cell.number_format = '[$₹-4009] #,##0.00'
        net_pay_cell.font = bold_font
        net_pay_cell.border = thin_border
        net_pay_cell.fill = row_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't auto-fit day columns to keep them compact (58px is approx width 6 in openpyxl)
        is_day_col = False
        if col[0].column > 4 and col[0].column <= (4 + num_days):
            is_day_col = True
            
        if is_day_col:
            ws.column_dimensions[col_letter].width = 6
        else:
            for cell in col:
                if cell.row < 4:  # Skip title row
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save to buffer and stream back
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    file_month_name = calendar.month_name[month]
    filename = f"Attendance_Payroll_Matrix_{file_month_name}_{year}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ----------------- DAILY LOGS ROUTES (RBAC ENFORCED) -----------------
@app.route('/logs', methods=['GET'])
def get_logs():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    subs = get_subordinate_usernames(user)
    if subs is None:
        # Admins see everyone's logs
        logs = list(db.intern_logs.find({}).sort("date_logged", -1))
    else:
        # Others see own logs + logs of subordinates
        logs = list(db.intern_logs.find({"intern_name": {"$in": subs}}).sort("date_logged", -1))
    
    return jsonify(serialize(logs))


@app.route('/submit', methods=['POST'])
def submit_log():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    # Interns and Employees can only log for themselves. Admins/Managers can log on behalf of others.
    if user['role'] in ['Intern', 'Employee']:
        name = user['username']
    else:
        name = data.get('intern_name', user['username'])

    date = data.get('date_logged')
    check_in = data.get('check_in', '')
    check_out = data.get('check_out', '')
    hours = data.get('hours_worked')
    task_category = data.get('task_category', '')
    tasks = data.get('tasks_completed')
    deliverable = data.get('deliverable_completed', 'No')
    blockers = data.get('blockers', 'None')
    skills_used = data.get('skills_used', '')
    mood = data.get('mood', 5)
    notes = data.get('notes', '')

    if not date or hours is None or not tasks:
        return jsonify({"error": "Missing required fields (date_logged, hours_worked, tasks_completed)"}), 400

    date_err = validate_log_date(date)
    if date_err:
        return jsonify({"error": date_err}), 400

    try:
        hours = float(hours)
        mood = int(mood)
    except ValueError:
        return jsonify({"error": "Invalid numeric format"}), 400

    try:
        db.intern_logs.insert_one({
            "intern_name": name,
            "date_logged": date,
            "check_in": check_in,
            "check_out": check_out,
            "hours_worked": hours,
            "task_category": task_category,
            "tasks_completed": tasks,
            "deliverable_completed": deliverable,
            "blockers": blockers,
            "skills_used": skills_used,
            "mood": mood,
            "notes": notes
        })
    except Exception as e:
        return jsonify({"error": f"Submission failed: {str(e)}"}), 500

    return jsonify({"message": "Daily log submitted successfully!"}), 201


@app.route('/logs/<log_id>', methods=['PUT'])
def update_log(log_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        log = db.intern_logs.find_one({"_id": ObjectId(log_id)})
    except Exception:
        return jsonify({"error": "Invalid log ID format"}), 400

    if not log:
        return jsonify({"error": "Daily log not found"}), 404

    # Find owner role
    owner = db.users.find_one({"username": log['intern_name']})
    owner_role = owner['role'] if owner else 'Intern'

    # Enforce edit restrictions:
    # 1. Interns and Employees can only edit their own logs.
    # 2. Managers can edit their own logs or logs of Interns/Employees, but NOT other Managers or Admins.
    if user['role'] in ['Intern', 'Employee'] and log['intern_name'] != user['username']:
        return jsonify({"error": "Forbidden: You cannot modify this log"}), 403
    elif user['role'] == 'Manager':
        if owner_role in ['Manager', 'Admin'] and log['intern_name'] != user['username']:
            return jsonify({"error": "Forbidden: Only Admin can modify/review a Manager's log"}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    date = data.get('date_logged')
    check_in = data.get('check_in', '')
    check_out = data.get('check_out', '')
    hours = data.get('hours_worked')
    task_category = data.get('task_category', '')
    tasks = data.get('tasks_completed')
    deliverable = data.get('deliverable_completed', 'No')
    blockers = data.get('blockers', 'None')
    skills_used = data.get('skills_used', '')
    mood = data.get('mood')
    notes = data.get('notes', '')

    if not date or hours is None or not tasks:
        return jsonify({"error": "Missing required fields (date_logged, hours_worked, tasks_completed)"}), 400

    date_err = validate_log_date(date)
    if date_err:
        return jsonify({"error": date_err}), 400

    try:
        hours = float(hours)
        mood = int(mood) if mood is not None else 5
    except ValueError:
        return jsonify({"error": "Invalid numeric format"}), 400

    update_fields = {
        "date_logged": date,
        "check_in": check_in,
        "check_out": check_out,
        "hours_worked": hours,
        "task_category": task_category,
        "tasks_completed": tasks,
        "deliverable_completed": deliverable,
        "blockers": blockers,
        "skills_used": skills_used,
        "mood": mood,
        "notes": notes
    }

    # If Admin/Manager is editing, they can also update the intern_name if it was supplied
    if user['role'] not in ['Intern', 'Employee'] and 'intern_name' in data:
        update_fields["intern_name"] = data.get('intern_name')

    db.intern_logs.update_one({"_id": ObjectId(log_id)}, {"$set": update_fields})
    return jsonify({"message": "Daily log updated successfully!"})


@app.route('/logs/<log_id>', methods=['DELETE'])
def delete_log(log_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        log = db.intern_logs.find_one({"_id": ObjectId(log_id)})
    except Exception:
        return jsonify({"error": "Invalid log ID format"}), 400

    if not log:
        return jsonify({"error": "Daily log not found"}), 404

    # Find owner role
    owner = db.users.find_one({"username": log['intern_name']})
    owner_role = owner['role'] if owner else 'Intern'

    # Enforce delete restrictions:
    # 1. Interns/Employees can only delete their own logs.
    # 2. Managers can delete their own logs or logs of Interns/Employees, but NOT other Managers or Admins.
    if user['role'] in ['Intern', 'Employee'] and log['intern_name'] != user['username']:
        return jsonify({"error": "Forbidden: You cannot delete this log"}), 403
    elif user['role'] == 'Manager':
        if owner_role in ['Manager', 'Admin'] and log['intern_name'] != user['username']:
            return jsonify({"error": "Forbidden: Only Admin can delete a Manager's log"}), 403

    db.intern_logs.delete_one({"_id": ObjectId(log_id)})
    return jsonify({"message": "Daily log deleted successfully!"})


# ----------------- TASKS ROUTES (RBAC ENFORCED) -----------------
@app.route('/tasks', methods=['GET'])
def get_tasks():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    subs = get_subordinate_usernames(user)
    if subs is None:
        # Admins see everyone's tasks
        tasks = list(db.tasks.find({}).sort("due_date", 1))
    else:
        # Others see own tasks + tasks of subordinates
        tasks = list(db.tasks.find({"intern_name": {"$in": subs}}).sort("due_date", 1))
        
    return jsonify(serialize(tasks))


@app.route('/tasks', methods=['POST'])
def add_task():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    # Validation of target user and role permissions
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    target_username = data.get('intern_name') # Target user to assign task
    task_name = data.get('task_name')
    category = data.get('category', '')
    assigned_date = data.get('assigned_date')
    due_date = data.get('due_date', '')
    priority = data.get('priority', 'Medium')
    status = data.get('status', 'Not Started')
    percent_done = data.get('percent_done', 0)
    notes = data.get('notes', '')

    if not target_username or not task_name or not assigned_date:
        return jsonify({"error": "Missing required fields (intern_name, task_name, assigned_date)"}), 400

    due_date_err = validate_future_or_today_date(due_date, "Due date")
    if due_date_err:
        return jsonify({"error": due_date_err}), 400

    target_user = db.users.find_one({"username": target_username})
    if not target_user:
        return jsonify({"error": "Target user not found"}), 404
        
    target_role = target_user['role']
    
    # Enforce role rules:
    # 1. Intern can only assign tasks to themselves
    if user['role'] == 'Intern' and target_username != user['username']:
        return jsonify({"error": "Forbidden: Interns can only assign tasks to themselves"}), 403

    # 2. Employee can only assign tasks to themselves or Interns
    if user['role'] == 'Employee' and target_username != user['username'] and target_role != 'Intern':
        return jsonify({"error": "Forbidden: Employees can only assign tasks to themselves or Interns"}), 403

    # 3. Managers and Admins can only have tasks assigned to them by Admins
    if target_role in ['Manager', 'Admin'] and user['role'] != 'Admin':
        return jsonify({"error": "Forbidden: Only Admin can assign tasks to Managers or Admins"}), 403

    try:
        db.tasks.insert_one({
            "intern_name": target_username,
            "task_name": task_name,
            "category": category,
            "assigned_date": assigned_date,
            "due_date": due_date,
            "priority": priority,
            "status": status,
            "percent_done": int(percent_done),
            "assigned_by": user['username'],
            "notes": notes,
            "started_at": None,
            "completed_at": None
        })
        
        # Create notification for target user (suppressed if self-assigned)
        if target_username != user['username']:
            try:
                db.notifications.insert_one({
                    "username": target_username,
                    "message": f"New task assigned by {user['full_name']}: '{task_name}'",
                    "type": "task_assignment",
                    "created_at": datetime.datetime.utcnow(),
                    "read": False
                })
            except Exception as ne:
                print(f"Failed to create notification: {ne}")
            
    except Exception as e:
        return jsonify({"error": f"Task creation failed: {str(e)}"}), 500


    return jsonify({"message": "Task created successfully!"}), 201


@app.route('/tasks/<task_id>', methods=['PUT'])
def update_task(task_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    status = data.get('status')
    percent_done = data.get('percent_done')
    completed_date = data.get('completed_date', '')
    notes = data.get('notes')

    try:
        task = db.tasks.find_one({"_id": ObjectId(task_id)})
    except Exception:
        return jsonify({"error": "Invalid task ID format"}), 400

    if not task:
        return jsonify({"error": "Task not found"}), 404
    
    # Access checks:
    # 1. Interns/Employees can only update tasks assigned to them
    # 2. Managers can only update tasks of Interns/Employees, or their own tasks
    owner = db.users.find_one({"username": task['intern_name']})
    owner_role = owner['role'] if owner else 'Intern'
    if user['role'] in ['Intern', 'Employee'] and task['intern_name'] != user['username']:
        return jsonify({"error": "Forbidden: You cannot modify this task"}), 403
    elif user['role'] == 'Manager':
        if owner_role in ['Manager', 'Admin'] and task['intern_name'] != user['username']:
            return jsonify({"error": "Forbidden: Only Admin can modify tasks of Managers or Admins"}), 403

    update_fields = {}
    if status is not None:
        update_fields["status"] = status
        
        # Track start time when moved to In Progress
        if status == 'In Progress' and not task.get('started_at'):
            update_fields["started_at"] = datetime.datetime.utcnow().isoformat() + 'Z'
            
        # Track completion time when moved to Completed
        if status == 'Completed':
            if not task.get('completed_at'):
                update_fields["completed_at"] = datetime.datetime.utcnow().isoformat() + 'Z'
            if not task.get('completed_date'):
                update_fields["completed_date"] = datetime.date.today().isoformat()
    if percent_done is not None:
        try:
            update_fields["percent_done"] = int(percent_done)
        except ValueError:
            pass
    if completed_date is not None:
        update_fields["completed_date"] = completed_date
    if notes is not None:
        update_fields["notes"] = notes

    if update_fields:
        db.tasks.update_one({"_id": ObjectId(task_id)}, {"$set": update_fields})

    return jsonify({"message": "Task updated successfully!"})


@app.route('/tasks/<task_id>', methods=['DELETE'])
def delete_task(task_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        task = db.tasks.find_one({"_id": ObjectId(task_id)})
    except Exception:
        return jsonify({"error": "Invalid task ID format"}), 400

    if not task:
        return jsonify({"error": "Task not found"}), 404
    
    # Interns cannot delete tasks.
    # Employees can only delete tasks if they created it.
    # Managers/Admins can delete any task, except tasks of other Managers/Admins (only Admin can).
    owner = db.users.find_one({"username": task['intern_name']})
    owner_role = owner['role'] if owner else 'Intern'
    if user['role'] == 'Intern':
        return jsonify({"error": "Forbidden"}), 403
    elif user['role'] == 'Employee' and task['assigned_by'] != user['username']:
        return jsonify({"error": "Forbidden: You can only delete tasks you created"}), 403
    elif user['role'] == 'Manager':
        if owner_role in ['Manager', 'Admin'] and task['intern_name'] != user['username']:
            return jsonify({"error": "Forbidden: Only Admin can delete tasks of Managers or Admins"}), 403

    db.tasks.delete_one({"_id": ObjectId(task_id)})
    return jsonify({"message": "Task deleted successfully!"})


# ----------------- SKILLS ROUTES (RBAC ENFORCED) -----------------
@app.route('/skills', methods=['GET'])
def get_skills():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    subs = get_subordinate_usernames(user)
    if subs is None:
        # Admins see everyone's skills
        skills = list(db.skills_log.find({}).sort("date_logged", -1))
    else:
        # Others see own skills + skills of subordinates
        skills = list(db.skills_log.find({"intern_name": {"$in": subs}}).sort("date_logged", -1))
        
    return jsonify(serialize(skills))


@app.route('/skills', methods=['POST'])
def add_skill():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    # Interns/Employees log skills for themselves.
    if user['role'] in ['Intern', 'Employee']:
        name = user['username']
    else:
        name = data.get('intern_name', user['username'])

    date = data.get('date_logged')
    skill_tool = data.get('skill_tool')
    category = data.get('category', '')
    resource = data.get('resource_course', '')
    hours = data.get('hours_spent')
    prof_before = data.get('proficiency_before')
    prof_after = data.get('proficiency_after')
    certificate = data.get('certificate', 'No')
    notes = data.get('notes', '')

    if not date or not skill_tool or hours is None:
        return jsonify({"error": "Missing required fields (date_logged, skill_tool, hours_spent)"}), 400

    try:
        hours = float(hours)
        prof_before = int(prof_before) if prof_before is not None else None
        prof_after = int(prof_after) if prof_after is not None else None
    except ValueError:
        return jsonify({"error": "Invalid format"}), 400

    try:
        db.skills_log.insert_one({
            "intern_name": name,
            "date_logged": date,
            "skill_tool": skill_tool,
            "category": category,
            "resource_course": resource,
            "hours_spent": hours,
            "proficiency_before": prof_before,
            "proficiency_after": prof_after,
            "certificate": certificate,
            "notes": notes
        })
    except Exception as e:
        return jsonify({"error": f"Skill logging failed: {str(e)}"}), 500

    return jsonify({"message": "Skill logged successfully!"}), 201


# ----------------- MENTOR FEEDBACK ROUTES (RBAC ENFORCED) -----------------
@app.route('/feedback', methods=['GET'])
def get_feedback():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    subs = get_subordinate_usernames(user)
    if subs is None:
        # Admins see everyone's feedback reviews
        feedback = list(db.mentor_feedback.find({}).sort("date_logged", -1))
    else:
        # Others see reviews written for/by them, or written for subordinates
        feedback = list(db.mentor_feedback.find({
            "$or": [
                {"intern_name": {"$in": subs}},
                {"feedback_from": user['username']}
            ]
        }).sort("date_logged", -1))
        
    return jsonify(serialize(feedback))


@app.route('/feedback', methods=['POST'])
def add_feedback():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    # Interns cannot log supervisor reviews.
    if user['role'] == 'Intern':
        return jsonify({"error": "Interns cannot log feedback reviews"}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    target_username = data.get('intern_name') # Target user to receive feedback
    date = data.get('date_logged')
    feedback_type = data.get('type')
    summary = data.get('feedback_summary')
    area_improve = data.get('area_to_improve', '')
    strength = data.get('strength_noted', '')
    action = data.get('action_taken', '')
    follow_up = data.get('follow_up', 'No')
    follow_up_date = data.get('follow_up_date', '')

    if not target_username or not date or not feedback_type or not summary:
        return jsonify({"error": "Missing required fields (intern_name, date_logged, type, feedback_summary)"}), 400

    follow_up_err = validate_future_or_today_date(follow_up_date, "Follow-up date")
    if follow_up_err:
        return jsonify({"error": follow_up_err}), 400

    target_user = db.users.find_one({"username": target_username})
    if not target_user:
        return jsonify({"error": "Target user not found"}), 404
        
    target_role = target_user['role']

    # Employees can ONLY leave feedback for Interns
    if user['role'] == 'Employee' and target_role != 'Intern':
        return jsonify({"error": "Employees can only log feedback for Interns"}), 403

    # Managers and Admins can only receive feedback from Admins
    if target_role in ['Manager', 'Admin'] and user['role'] != 'Admin':
        return jsonify({"error": "Only Administrators can submit feedback for Managers or Admins"}), 403

    try:
        db.mentor_feedback.insert_one({
            "intern_name": target_username,
            "date_logged": date,
            "feedback_from": user['username'],
            "type": feedback_type,
            "feedback_summary": summary,
            "area_to_improve": area_improve,
            "strength_noted": strength,
            "action_taken": action,
            "follow_up": follow_up,
            "follow_up_date": follow_up_date
        })
    except Exception as e:
        return jsonify({"error": f"Feedback submission failed: {str(e)}"}), 500

    return jsonify({"message": "Mentor feedback logged successfully!"}), 201


# ----------------- NOTIFICATIONS API ENDPOINTS -----------------
@app.route('/notifications', methods=['GET'])
def get_notifications():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    notifications = list(db.notifications.find({"username": user['username']}).sort("created_at", -1).limit(20))
    return jsonify(serialize(notifications))


@app.route('/notifications/read', methods=['POST'])
def mark_all_notifications_read():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    db.notifications.update_many(
        {"username": user['username'], "read": False},
        {"$set": {"read": True}}
    )
    return jsonify({"message": "All notifications marked as read"})


@app.route('/notifications/<notif_id>/read', methods=['POST'])
def mark_notification_read(notif_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        db.notifications.update_one(
            {"_id": ObjectId(notif_id), "username": user['username']},
            {"$set": {"read": True}}
        )
    except Exception:
        return jsonify({"error": "Invalid notification ID"}), 400
        
    return jsonify({"message": "Notification marked as read"})


# ----------------- USER MANAGEMENT: UPDATE ROLE (Admin Only) -----------------
@app.route('/users/<user_id>/role', methods=['PUT'])
def update_user_role(user_id):
    user = get_logged_in_user()
    if not user or user['role'] != 'Admin':
        return jsonify({"error": "Unauthorized: Only Admins can modify roles"}), 401
    
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400
        
    new_role = data.get('role')
    if new_role not in ['Admin', 'Manager', 'Employee', 'Intern']:
        return jsonify({"error": "Invalid role"}), 400
        
    try:
        target_user = db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        return jsonify({"error": "Invalid user ID format"}), 400
        
    if not target_user:
        return jsonify({"error": "User not found"}), 404
        
    # Prevent changing the role of oneself (prevents admin locking themselves out)
    if str(target_user['_id']) == str(user['_id']):
        return jsonify({"error": "You cannot change your own role to prevent lockout"}), 400
        
    # Prevent changing the role of the primary admin
    if target_user['username'] == 'admin':
        return jsonify({"error": "You cannot change the role of the primary administrator"}), 400
        
    db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"role": new_role}})
    return jsonify({"message": f"User {target_user['full_name']} role updated to {new_role}!"})


# ----------------- USER MANAGEMENT: APPROVE USER (Admin Only) -----------------
@app.route('/users/<user_id>/approve', methods=['PUT'])
def approve_user(user_id):
    user = get_logged_in_user()
    if not user or user['role'] != 'Admin':
        return jsonify({"error": "Unauthorized: Only Admins can approve users"}), 401
        
    try:
        target_user = db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        return jsonify({"error": "Invalid user ID format"}), 400
        
    if not target_user:
        return jsonify({"error": "User not found"}), 404
        
    db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"approved": True}})
    
    # Create notification for target user
    try:
        db.notifications.insert_one({
            "username": target_user['username'],
            "message": "Your account has been approved by the Admin! You can now log in.",
            "type": "account_approved",
            "created_at": datetime.datetime.utcnow(),
            "read": False
        })
    except Exception as ne:
        print(f"Failed to create approval notification: {ne}")
        
    return jsonify({"message": f"User {target_user['full_name']} has been approved successfully!"})


# ----------------- USER MANAGEMENT: RESTRICT USER (Admin & Manager) -----------------
@app.route('/users/<user_id>/restrict', methods=['PUT'])
def restrict_user(user_id):
    user = get_logged_in_user()
    if not user or user['role'] not in ['Admin', 'Manager']:
        return jsonify({"error": "Unauthorized: Access denied"}), 401
        
    try:
        target_user = db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        return jsonify({"error": "Invalid user ID format"}), 400
        
    if not target_user:
        return jsonify({"error": "User not found"}), 404
        
    # Prevent self-restriction
    if str(target_user['_id']) == str(user['_id']):
        return jsonify({"error": "You cannot restrict yourself"}), 400
        
    role_weights = {"Intern": 1, "Employee": 2, "Manager": 3, "Admin": 4}
    user_weight = role_weights.get(user['role'], 0)
    target_weight = role_weights.get(target_user['role'], 0)
    
    # Enforce hierarchical access control: Rank(A) > Rank(B) (Bypassed for Admin)
    if user['role'] != 'Admin' and user_weight <= target_weight:
        return jsonify({"error": "Forbidden: You can only restrict users with a lower rank than yours"}), 403
        
    db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"restricted": True}})
    
    # Create notification for target user
    try:
        db.notifications.insert_one({
            "username": target_user['username'],
            "message": "Your account has been temporarily restricted by an administrator or manager. Please contact support.",
            "type": "account_restricted",
            "created_at": datetime.datetime.utcnow(),
            "read": False
        })
    except Exception as ne:
        print(f"Failed to create restriction notification: {ne}")
        
    return jsonify({"message": f"User {target_user['full_name']} has been restricted successfully!"})


# ----------------- USER MANAGEMENT: UNRESTRICT USER (Admin & Manager) -----------------
@app.route('/users/<user_id>/unrestrict', methods=['PUT'])
def unrestrict_user(user_id):
    user = get_logged_in_user()
    if not user or user['role'] not in ['Admin', 'Manager']:
        return jsonify({"error": "Unauthorized: Access denied"}), 401
        
    try:
        target_user = db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        return jsonify({"error": "Invalid user ID format"}), 400
        
    if not target_user:
        return jsonify({"error": "User not found"}), 404
        
    role_weights = {"Intern": 1, "Employee": 2, "Manager": 3, "Admin": 4}
    user_weight = role_weights.get(user['role'], 0)
    target_weight = role_weights.get(target_user['role'], 0)
    
    # Enforce hierarchical access control: Rank(A) > Rank(B) (Bypassed for Admin)
    if user['role'] != 'Admin' and user_weight <= target_weight:
        return jsonify({"error": "Forbidden: You can only unrestrict users with a lower rank than yours"}), 403
        
    db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"restricted": False}})
    
    # Create notification for target user
    try:
        db.notifications.insert_one({
            "username": target_user['username'],
            "message": "Your restriction has been lifted! You can now log in and log hours again.",
            "type": "account_unrestricted",
            "created_at": datetime.datetime.utcnow(),
            "read": False
        })
    except Exception as ne:
        print(f"Failed to create unrestriction notification: {ne}")
        
    return jsonify({"message": f"User {target_user['full_name']} restriction lifted successfully!"})


# ----------------- SESSION LOGS API ENDPOINT (Admin & Manager Only) -----------------
@app.route('/session_logs', methods=['GET'])
def get_session_logs():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if user['role'] not in ['Admin', 'Manager', 'Guest']:
        return jsonify({"error": "Forbidden: Access restricted"}), 403

    try:
        # Enforce hierarchical access control:
        # Admin can view all login/logout session logs
        if user['role'] in ['Admin', 'Guest']:
            logs = list(db.session_logs.find({}).sort("login_time", -1))
        # Manager can view Employee and Intern logs, plus their own
        else:
            sub_users = list(db.users.find({"role": {"$in": ["Employee", "Intern"]}}, {"username": 1}))
            sub_usernames = [u['username'] for u in sub_users]
            sub_usernames.append(user['username'])
            
            logs = list(db.session_logs.find({"username": {"$in": sub_usernames}}).sort("login_time", -1))
            
        return jsonify(serialize(logs))
    except Exception as e:
        return jsonify({"error": f"Failed to fetch session logs: {str(e)}"}), 500


@app.route('/session_logs/edit', methods=['PUT'])
def edit_session_log():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if user['role'] != 'Admin':
        return jsonify({"error": "Forbidden: Only Admins can edit session logs"}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    target_username = data.get('username')
    target_date = data.get('date')
    new_login = data.get('login_time')
    new_logout = data.get('logout_time')

    if not target_username or not target_date or not new_login or not new_logout:
        return jsonify({"error": "Missing required fields"}), 400

    try:
        try:
            dt_login = datetime.datetime.fromisoformat(new_login.replace('Z', '+00:00')).replace(tzinfo=None)
            dt_logout = datetime.datetime.fromisoformat(new_logout.replace('Z', '+00:00')).replace(tzinfo=None)
        except ValueError:
            return jsonify({"error": "Invalid date format"}), 400

        start_of_day = dt_login.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + datetime.timedelta(days=1)
        
        db.session_logs.delete_many({
            "username": target_username,
            "login_time": {"$gte": start_of_day, "$lt": end_of_day}
        })
        
        target_user = db.users.find_one({"username": target_username})
        full_name = target_user['full_name'] if target_user else target_username
        role = target_user['role'] if target_user else "Intern"

        db.session_logs.insert_one({
            "username": target_username,
            "full_name": full_name,
            "role": role,
            "login_time": dt_login,
            "logout_time": dt_logout
        })

        return jsonify({"message": "Session log updated successfully"})

    except Exception as e:
        return jsonify({"error": f"Failed to update session log: {str(e)}"}), 500



# ----------------- COLLABORATION: MEDIA UPLOAD -----------------
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/uploads/<filename>')
def serve_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/upload', methods=['POST'])
def upload_file():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected for uploading"}), 400

    if file:
        import uuid
        from werkzeug.utils import secure_filename
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        filename = f"{uuid.uuid4()}.{ext}" if ext else str(uuid.uuid4())
        filename = secure_filename(filename)
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(file_path)

        file_url = f"/uploads/{filename}"
        return jsonify({"url": file_url}), 200
    
    return jsonify({"error": "Upload failed"}), 400


# ----------------- COLLABORATION: PUBLIC KEYS FOR E2EE -----------------
@app.route('/users/public_key', methods=['PUT'])
def update_public_key():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    if not data or 'public_key' not in data:
        return jsonify({"error": "Missing public_key in request"}), 400

    db.users.update_one({"username": user["username"]}, {"$set": {"public_key": data["public_key"]}})
    return jsonify({"message": "Public key updated successfully!"}), 200

@app.route('/users/<username>/public_key', methods=['GET'])
def get_user_public_key(username):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    target = db.users.find_one({"username": username}, {"public_key": 1, "full_name": 1})
    if not target:
        return jsonify({"error": "User not found"}), 404

    if 'public_key' not in target:
        return jsonify({"error": f"User {target.get('full_name')} has not activated secure messaging yet (no public key)"}), 404

    return jsonify({"public_key": target["public_key"]}), 200


# ----------------- COLLABORATION: ANNOUNCEMENTS -----------------
@app.route('/announcements', methods=['GET'])
def get_announcements():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        announcements = list(db.announcements.find({}).sort("created_at", -1).limit(50))
        announcements.reverse()
        return jsonify(serialize(announcements))
    except Exception as e:
        return jsonify({"error": f"Failed to fetch announcements: {str(e)}"}), 500

@app.route('/announcements', methods=['POST'])
def add_announcement():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    if not data or 'content' not in data:
        return jsonify({"error": "Missing content in request"}), 400

    content = data.get('content')
    media_url = data.get('media_url', '')
    media_type = data.get('media_type', '')

    try:
        db.announcements.insert_one({
            "sender": user["username"],
            "sender_fullname": user["full_name"],
            "sender_role": user["role"],
            "content": content,
            "media_url": media_url,
            "media_type": media_type,
            "created_at": datetime.datetime.utcnow()
        })

        # Create notifications for all other approved active users
        other_users = db.users.find({"username": {"$ne": user["username"]}, "approved": True})
        for u in other_users:
            try:
                snippet = content[:40] + "..." if len(content) > 40 else content
                db.notifications.insert_one({
                    "username": u["username"],
                    "message": f"New announcement from {user['full_name']}: '{snippet}'",
                    "type": "announcement",
                    "created_at": datetime.datetime.utcnow(),
                    "read": False
                })
            except Exception as ne:
                print(f"Failed to create announcement notification: {ne}")

        return jsonify({"message": "Announcement posted successfully!"}), 201
    except Exception as e:
        return jsonify({"error": f"Failed to post announcement: {str(e)}"}), 500


# ----------------- COLLABORATION: SECURE ONE-TO-ONE MESSAGING (E2EE) -----------------
@app.route('/direct_messages', methods=['GET'])
def get_direct_messages():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    recipient = request.args.get('recipient')
    if not recipient:
        return jsonify({"error": "Missing recipient parameter"}), 400

    try:
        # Mark incoming messages from recipient to current user as seen
        db.direct_messages.update_many(
            {"sender": recipient, "recipient": user["username"], "seen": {"$ne": True}},
            {"$set": {"seen": True}}
        )

        query = {
            "$or": [
                {"sender": user["username"], "recipient": recipient},
                {"sender": recipient, "recipient": user["username"]}
            ]
        }
        messages = list(db.direct_messages.find(query).sort("created_at", 1))
        return jsonify(serialize(messages))
    except Exception as e:
        return jsonify({"error": f"Failed to fetch messages: {str(e)}"}), 500

@app.route('/direct_messages', methods=['POST'])
def add_direct_message():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    if not data or not all(k in data for k in ['recipient', 'ciphertext', 'iv', 'sender_key_enc', 'recipient_key_enc']):
        return jsonify({"error": "Missing required encrypted payload fields"}), 400

    recipient = data.get('recipient')
    ciphertext = data.get('ciphertext')
    iv = data.get('iv')
    sender_key_enc = data.get('sender_key_enc')
    recipient_key_enc = data.get('recipient_key_enc')

    target_user = db.users.find_one({"username": recipient})
    if not target_user:
        return jsonify({"error": "Recipient not found"}), 404

    try:
        db.direct_messages.insert_one({
            "sender": user["username"],
            "recipient": recipient,
            "ciphertext": ciphertext,
            "iv": iv,
            "sender_key_enc": sender_key_enc,
            "recipient_key_enc": recipient_key_enc,
            "created_at": datetime.datetime.utcnow(),
            "seen": False
        })

        # Create message notification for target user
        try:
            db.notifications.insert_one({
                "username": recipient,
                "message": f"New secure message from {user['full_name']}",
                "type": "direct_message",
                "created_at": datetime.datetime.utcnow(),
                "read": False
            })
        except Exception as ne:
            print(f"Failed to create message notification: {ne}")

        return jsonify({"message": "Message sent successfully!"}), 201
    except Exception as e:
        return jsonify({"error": f"Failed to send message: {str(e)}"}), 500


# ----------------- COLLABORATION: CONVERSATIONS LIST (E2EE) -----------------
@app.route('/direct_messages/conversations', methods=['GET'])
def get_conversations():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        # Find all DMs involving the current user
        query = {"$or": [{"sender": user["username"]}, {"recipient": user["username"]}]}
        messages = list(db.direct_messages.find(query, {"sender": 1, "recipient": 1, "created_at": 1, "seen": 1}))
        
        conversations = {}
        for m in messages:
            other = m["recipient"] if m["sender"] == user["username"] else m["sender"]
            if other not in conversations:
                conversations[other] = {
                    "count": 0,
                    "unread_count": 0,
                    "received_count": 0,
                    "last_message_time": m["created_at"]
                }
            
            # Total messages in the conversation
            conversations[other]["count"] += 1
            
            # Messages sent BY THE OTHER USER
            if m["sender"] == other:
                conversations[other]["received_count"] += 1
                if not m.get("seen", False):
                    conversations[other]["unread_count"] += 1

            if m["created_at"] > conversations[other]["last_message_time"]:
                conversations[other]["last_message_time"] = m["created_at"]
        
        return jsonify(conversations)
    except Exception as e:
        return jsonify({"error": f"Failed to fetch conversation stats: {str(e)}"}), 500


# ----------------- DELETE MESSAGE & ANNOUNCEMENT ENDPOINTS -----------------
@app.route('/announcements/<announcement_id>', methods=['DELETE'])
def delete_announcement(announcement_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        ann = db.announcements.find_one({"_id": ObjectId(announcement_id)})
        if not ann:
            return jsonify({"error": "Announcement not found"}), 404
        
        # Sender, Admin, or Manager can delete
        if ann["sender"] == user["username"] or user["role"] in ["Admin", "Manager"]:
            db.announcements.delete_one({"_id": ObjectId(announcement_id)})
            
            # Clean up notifications related to this announcement
            db.notifications.delete_many({
                "type": "announcement",
                "message": {"$regex": f"New announcement from {ann['sender_fullname']}"}
            })
            
            return jsonify({"message": "Announcement deleted successfully!"}), 200
        else:
            return jsonify({"error": "Permission denied. You can only delete your own posts."}), 403
    except Exception as e:
        return jsonify({"error": f"Failed to delete announcement: {str(e)}"}), 500

@app.route('/direct_messages/<message_id>', methods=['DELETE'])
def delete_direct_message(message_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        msg = db.direct_messages.find_one({"_id": ObjectId(message_id)})
        if not msg:
            return jsonify({"error": "Message not found"}), 404
        
        # Only the sender can delete their own message (unsend)
        if msg["sender"] == user["username"]:
            db.direct_messages.delete_one({"_id": ObjectId(message_id)})
            
            # Clean up unread notifications related to this DM from sender to recipient
            db.notifications.delete_many({
                "username": msg["recipient"],
                "type": "direct_message",
                "message": f"New secure message from {user['full_name']}",
                "read": False
            })
            
            return jsonify({"message": "Message deleted successfully!"}), 200
        else:
            return jsonify({"error": "Permission denied. You can only delete messages you sent."}), 403
    except Exception as e:
        return jsonify({"error": f"Failed to delete message: {str(e)}"}), 500


# Helper to ensure leave balances are initialized for a user
def ensure_leave_balances(username):
    current_year = datetime.datetime.utcnow().year
    # Total allocations: CL=12 days, SL=12 days, hours=96 hours (8 hours/month * 12)
    db.leave_balances.update_one(
        {"username": username, "year": current_year},
        {"$setOnInsert": {
            "cl_used": 0.0,
            "sl_used": 0.0,
            "hours_used": 0.0
        }},
        upsert=True
    )

@app.route('/leaves/balances', methods=['GET'])
def get_leave_balances():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    
    current_year = datetime.datetime.utcnow().year
    ensure_leave_balances(user['username'])
    bal = db.leave_balances.find_one({"username": user['username'], "year": current_year})
    
    return jsonify({
        "cl_used": bal.get("cl_used", 0.0),
        "sl_used": bal.get("sl_used", 0.0),
        "hours_used": bal.get("hours_used", 0.0),
        "cl_allocation": 12.0,
        "sl_allocation": 12.0,
        "hours_allocation": 96.0,
        "year": current_year
    })

@app.route('/leaves/apply', methods=['POST'])
def apply_leave():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if user['role'] not in ['Manager', 'Employee']:
        return jsonify({"error": "Forbidden: Only Managers and Employees can apply for leaves."}), 403

    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    leave_type = data.get('leave_type')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    reason = data.get('reason', '')
    days_requested = float(data.get('days_requested', 0.0))
    hours_requested = float(data.get('hours_requested', 0.0))
    doc_note_url = data.get('doc_note_url', None)

    if not leave_type or not start_date or not end_date or not reason:
        return jsonify({"error": "Missing required fields (leave_type, start_date, end_date, reason)"}), 400

    if start_date > end_date:
        return jsonify({"error": "Start date cannot be after end date."}), 400

    if leave_type == 'Hours':
        if hours_requested <= 0:
            return jsonify({"error": "Hours requested must be greater than 0."}), 400
    else:
        if days_requested <= 0:
            return jsonify({"error": "Days requested must be greater than 0."}), 400

    ensure_leave_balances(user['username'])
    current_year = datetime.datetime.utcnow().year
    bal = db.leave_balances.find_one({"username": user['username'], "year": current_year})

    # Validate balances unless it is Loss of Pay (LOP)
    if leave_type != 'LOP':
        if leave_type == 'CL':
            cl_remaining = 12.0 - bal.get("cl_used", 0.0)
            if days_requested > cl_remaining:
                return jsonify({"error": f"Insufficient Casual Leave balance. Remaining: {cl_remaining} days, Requested: {days_requested} days."}), 400
        elif leave_type == 'SL':
            sl_remaining = 12.0 - bal.get("sl_used", 0.0)
            if days_requested > sl_remaining:
                return jsonify({"error": f"Insufficient Sick Leave balance. Remaining: {sl_remaining} days, Requested: {days_requested} days."}), 400
        elif leave_type == 'Hours':
            hours_remaining = 96.0 - bal.get("hours_used", 0.0)
            if hours_requested > hours_remaining:
                return jsonify({"error": f"Insufficient Timing Leave balance. Remaining: {hours_remaining} hours, Requested: {hours_requested} hours."}), 400

    # Create request document
    req = {
        "username": user['username'],
        "full_name": user['full_name'],
        "role": user['role'],
        "leave_type": leave_type,
        "start_date": start_date,
        "end_date": end_date,
        "reason": reason,
        "days_requested": days_requested,
        "hours_requested": hours_requested,
        "doc_note_url": doc_note_url,
        "status": "Pending",
        "submitted_at": datetime.datetime.utcnow().isoformat() + 'Z',
        "actioned_by": None,
        "actioned_at": None
    }

    db.leave_requests.insert_one(req)

    # Notify Admins and Managers about the new leave request
    try:
        # Notify all Admins
        admins = list(db.users.find({"role": "Admin"}))
        for admin in admins:
            db.notifications.insert_one({
                "username": admin['username'],
                "message": f"New leave request ({leave_type}) submitted by {user['full_name']} ({user['role']}).",
                "type": "leave_request_alert",
                "created_at": datetime.datetime.utcnow(),
                "read": False
            })
        
        # If the requester is an Employee, notify all Managers
        if user['role'] == 'Employee':
            managers = list(db.users.find({"role": "Manager"}))
            for manager in managers:
                db.notifications.insert_one({
                    "username": manager['username'],
                    "message": f"New leave request ({leave_type}) submitted by {user['full_name']} (Employee).",
                    "type": "leave_request_alert",
                    "created_at": datetime.datetime.utcnow(),
                    "read": False
                })
    except Exception as e:
        print(f"Failed to create leave request notifications: {e}")

    return jsonify({"message": "Leave request submitted successfully!"}), 201

@app.route('/leaves/requests', methods=['GET'])
def get_leave_requests():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        # Enforce role hierarchy for leave requests visibility
        if user['role'] == 'Admin':
            # Admin sees all leave requests
            requests = list(db.leave_requests.find({}).sort("submitted_at", -1))
        elif user['role'] == 'Manager':
            # Manager sees subordinate Employee requests + their own requests
            sub_users = list(db.users.find({"role": "Employee"}, {"username": 1}))
            sub_usernames = [u['username'] for u in sub_users]
            sub_usernames.append(user['username'])
            requests = list(db.leave_requests.find({"username": {"$in": sub_usernames}}).sort("submitted_at", -1))
        elif user['role'] == 'Employee':
            # Employee sees their own requests
            requests = list(db.leave_requests.find({"username": user['username']}).sort("submitted_at", -1))
        else:
            # Interns or other roles see no leave requests
            requests = []
        
        return jsonify(serialize(requests))
    except Exception as e:
        return jsonify({"error": f"Failed to fetch requests: {str(e)}"}), 500

@app.route('/leaves/action/<request_id>', methods=['POST'])
def action_leave_request(request_id):
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if user['role'] not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden: Only Admin or Manager can action requests"}), 403

    data = request.json
    if not data or 'action' not in data:
        return jsonify({"error": "Missing action parameter"}), 400

    action = data['action'] # 'Approve' or 'Reject'
    if action not in ['Approve', 'Reject']:
        return jsonify({"error": "Invalid action value"}), 400

    try:
        req = db.leave_requests.find_one({"_id": ObjectId(request_id)})
    except Exception:
        return jsonify({"error": "Invalid request ID format"}), 400

    if not req:
        return jsonify({"error": "Leave request not found"}), 404

    if req['status'] != 'Pending':
        return jsonify({"error": "Leave request has already been actioned"}), 400

    # Manager cannot action Admin requests or other Manager requests
    if user['role'] == 'Manager':
        target_user = db.users.find_one({"username": req['username']})
        target_role = target_user['role'] if target_user else 'Intern'
        if target_role in ['Admin', 'Manager']:
            return jsonify({"error": "Forbidden: Managers cannot action requests from other Managers or Admins"}), 403

    status_val = 'Approved' if action == 'Approve' else 'Rejected'

    # If approved, deduct leave balance
    if action == 'Approve' and req['leave_type'] != 'LOP':
        target_username = req['username']
        ensure_leave_balances(target_username)
        
        # Deduct
        update_query = {}
        if req['leave_type'] == 'CL':
            update_query = {"$inc": {"cl_used": req['days_requested']}}
        elif req['leave_type'] == 'SL':
            update_query = {"$inc": {"sl_used": req['days_requested']}}
        elif req['leave_type'] == 'Hours':
            update_query = {"$inc": {"hours_used": req['hours_requested']}}

        if update_query:
            current_year = datetime.datetime.utcnow().year
            db.leave_balances.update_one({"username": target_username, "year": current_year}, update_query)

    # Update request status
    db.leave_requests.update_one(
        {"_id": ObjectId(request_id)},
        {"$set": {
            "status": status_val,
            "actioned_by": user['username'],
            "actioned_fullname": user['full_name'],
            "actioned_at": datetime.datetime.utcnow().isoformat() + 'Z'
        }}
    )

    # Trigger notification back to the employee
    try:
        db.notifications.insert_one({
            "username": req['username'],
            "message": f"Your leave request for {req['leave_type']} ({req['start_date']} to {req['end_date']}) has been {status_val} by {user['full_name']}.",
            "type": "leave_status",
            "created_at": datetime.datetime.utcnow(),
            "read": False
        })
    except Exception as ne:
        print(f"Failed to create leave notification: {ne}")

    return jsonify({"message": f"Leave request successfully {status_val}!"})

@app.route('/leaves/team-balances', methods=['GET'])
def get_team_leave_balances():
    user = get_logged_in_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    if user['role'] not in ['Admin', 'Manager']:
        return jsonify({"error": "Forbidden"}), 403

    try:
        # Enforce hierarchy rules:
        # Managers see Employees only
        # Admins see Managers & Employees
        if user['role'] == 'Admin':
            users_list = list(db.users.find({"role": {"$in": ["Manager", "Employee"]}}, {"username": 1, "full_name": 1, "role": 1, "title": 1}))
        else:
            users_list = list(db.users.find({"role": "Employee"}, {"username": 1, "full_name": 1, "role": 1, "title": 1}))

        # Filter out test users from list
        users_list = [u for u in users_list if not (
            'test' in u['username'].lower() or 
            'test' in u.get('full_name', '').lower() or 
            'pending' in u.get('full_name', '').lower() or
            u['username'].startswith('mgr_') or 
            u['username'].startswith('emp_') or 
            u['username'].startswith('int_')
        )]

        current_year = datetime.datetime.utcnow().year
        team_balances = []
        for u in users_list:
            ensure_leave_balances(u['username'])
            bal = db.leave_balances.find_one({"username": u['username'], "year": current_year})
            team_balances.append({
                "username": u['username'],
                "full_name": u['full_name'],
                "role": u['role'],
                "title": u.get("title", ""),
                "cl_used": bal.get("cl_used", 0.0),
                "sl_used": bal.get("sl_used", 0.0),
                "hours_used": bal.get("hours_used", 0.0),
                "cl_allocation": 12.0,
                "sl_allocation": 12.0,
                "hours_allocation": 96.0
            })
        
        return jsonify(team_balances)
    except Exception as e:
        return jsonify({"error": f"Failed to fetch team leave balances: {str(e)}"}), 500


HOSTING_EXCEL_PATH = r"C:\Users\s.anirudh\Downloads\Domain & Hosting Account Details.xlsx"

def import_hosting_from_excel_if_empty():
    try:
        if db.hosting_details.count_documents({}) > 0:
            return
    except Exception:
        return
    
    excel_path = HOSTING_EXCEL_PATH
    if not os.path.exists(excel_path):
        excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Domain & Hosting Account Details.xlsx")
        
    if not os.path.exists(excel_path):
        return
        
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]
            while headers and headers[-1].startswith('Col') and headers[-1] != 'Col0':
                headers.pop()
                
            for r_idx, r in enumerate(rows[1:], start=2):
                if all(v is None for v in r):
                    continue
                doc = {
                    "category": sheet_name,
                    "created_at": datetime.datetime.utcnow()
                }
                for c_idx, val in enumerate(r):
                    if c_idx < len(headers):
                        header = headers[c_idx]
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            val = val.strftime("%Y-%m-%d")
                        doc[header] = val if val is not None else ""
                db.hosting_details.insert_one(doc)
        print("Successfully imported hosting details from Excel to MongoDB!")
    except Exception as e:
        print(f"Failed to import hosting details: {e}")

@app.route('/admin/hosting-details', methods=['GET'])
def get_hosting_details_route():
    user = get_logged_in_user()
    if not user or user['role'] != 'Admin':
        return jsonify({"error": "Unauthorized"}), 401

    try:
        import_hosting_from_excel_if_empty()
        
        categories = ['GoDaddy', 'CPanel Domain', 'Domain With Us', 'Domain With Client']
        data = {cat: [] for cat in categories}
        
        docs = list(db.hosting_details.find({}))
        # Sort manually by Sl No if it is numeric
        def get_sl_no(d):
            try:
                return int(d.get('Sl No', 9999))
            except (ValueError, TypeError):
                return 9999

        docs.sort(key=get_sl_no)

        for doc in docs:
            cat = doc.get('category', 'GoDaddy')
            if cat not in data:
                data[cat] = []
            
            item = {
                "id": str(doc['_id'])
            }
            for k, v in doc.items():
                if k not in ['_id', 'category', 'created_at']:
                    item[k] = v
            data[cat].append(item)
            
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": f"Failed to read hosting details: {str(e)}"}), 500

@app.route('/admin/hosting-details/add', methods=['POST'])
def add_hosting_detail_route():
    user = get_logged_in_user()
    if not user or user['role'] != 'Admin':
        return jsonify({"error": "Unauthorized"}), 401

    try:
        req_data = request.get_json() or {}
        category = req_data.get('sheet_name')
        new_values = req_data.get('values', {})

        if not category:
            return jsonify({"error": "Category is required"}), 400

        last_item = db.hosting_details.find_one({"category": category}, sort=[("Sl No", -1)])
        sl_no = 1
        if last_item and last_item.get('Sl No'):
            try:
                sl_no = int(last_item['Sl No']) + 1
            except (ValueError, TypeError):
                pass

        doc = {
            "category": category,
            "created_at": datetime.datetime.utcnow(),
            "Sl No": sl_no
        }
        for k, v in new_values.items():
            if k not in ['id', 'category', 'created_at', 'Sl No']:
                doc[k] = v

        db.hosting_details.insert_one(doc)
        return jsonify({"message": "Hosting detail added successfully"})
    except Exception as e:
        return jsonify({"error": f"Failed to add hosting detail: {str(e)}"}), 500

@app.route('/admin/hosting-details/edit', methods=['PUT'])
def edit_hosting_detail_route():
    user = get_logged_in_user()
    if not user or user['role'] != 'Admin':
        return jsonify({"error": "Unauthorized"}), 401

    try:
        req_data = request.get_json() or {}
        item_id = req_data.get('row_idx')
        updated_values = req_data.get('values', {})

        if not item_id:
            return jsonify({"error": "Item ID is required"}), 400

        update_fields = {}
        for k, v in updated_values.items():
            if k not in ['id', 'category', 'created_at', 'Sl No']:
                update_fields[k] = v

        db.hosting_details.update_one({"_id": ObjectId(item_id)}, {"$set": update_fields})
        return jsonify({"message": "Hosting detail updated successfully"})
    except Exception as e:
        return jsonify({"error": f"Failed to update hosting detail: {str(e)}"}), 500

@app.route('/admin/hosting-details/delete', methods=['DELETE'])
def delete_hosting_detail_route():
    user = get_logged_in_user()
    if not user or user['role'] != 'Admin':
        return jsonify({"error": "Unauthorized"}), 401

    try:
        req_data = request.get_json() or {}
        item_id = req_data.get('row_idx')

        if not item_id:
            return jsonify({"error": "Item ID is required"}), 400

        item = db.hosting_details.find_one({"_id": ObjectId(item_id)})
        if not item:
            return jsonify({"error": "Item not found"}), 404
        category = item['category']

        db.hosting_details.delete_one({"_id": ObjectId(item_id)})
        
        category_items = list(db.hosting_details.find({"category": category}).sort("created_at", 1))
        for idx, cat_item in enumerate(category_items, start=1):
            db.hosting_details.update_one({"_id": cat_item['_id']}, {"$set": {"Sl No": idx}})

        return jsonify({"message": "Hosting detail deleted successfully"})
    except Exception as e:
        return jsonify({"error": f"Failed to delete hosting detail: {str(e)}"}), 500

@app.route('/admin/hosting-details/export', methods=['GET'])
def export_hosting_details_route():
    user = get_logged_in_user()
    if not user or user['role'] != 'Admin':
        return jsonify({"error": "Unauthorized"}), 401

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        categories = ['GoDaddy', 'CPanel Domain', 'Domain With Us', 'Domain With Client']
        headers_map = {
            'GoDaddy': ['Sl No', 'Domains', 'Renewal Date', 'domain cost'],
            'CPanel Domain': ['Sl No', 'Domains', 'hosting renewal cost', 'date'],
            'Domain With Us': ['Sl No', 'Domains', 'Domain With Us or Client'],
            'Domain With Client': ['Sl No', 'Domains']
        }
        
        for cat in categories:
            sheet = wb.create_sheet(title=cat)
            headers = headers_map[cat]
            sheet.append(headers)
            
            docs = list(db.hosting_details.find({"category": cat}))
            def get_sl_no(d):
                try:
                    return int(d.get('Sl No', 9999))
                except (ValueError, TypeError):
                    return 9999
            docs.sort(key=get_sl_no)

            for doc in docs:
                row_data = []
                for h in headers:
                    row_data.append(doc.get(h, ""))
                sheet.append(row_data)
                
        temp_filename = "Domain_Hosting_Account_Details_Export.xlsx"
        wb.save(temp_filename)
        
        return send_from_directory(
            os.path.abspath(os.path.dirname(__file__)),
            temp_filename,
            as_attachment=True,
            download_name="Domain & Hosting Account Details.xlsx"
        )
    except Exception as e:
        return jsonify({"error": f"Failed to export hosting details: {str(e)}"}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=True, host='0.0.0.0', port=port)
