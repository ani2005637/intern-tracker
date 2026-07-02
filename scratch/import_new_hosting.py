import os
import sys
import datetime
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import db

NEW_EXCEL_PATH = r"C:\Users\s.anirudh\Downloads\tracker\Domain & Hosting Account Details (1).xlsx"

def import_new_hosting():
    if not os.path.exists(NEW_EXCEL_PATH):
        print(f"Error: Updated Excel file not found at {NEW_EXCEL_PATH}")
        return

    try:
        # Clear existing hosting details collection
        print("Clearing old domain and hosting details from MongoDB...")
        db.hosting_details.delete_many({})

        print(f"Reading updated Excel file: {NEW_EXCEL_PATH}...")
        wb = openpyxl.load_workbook(NEW_EXCEL_PATH, read_only=True)
        
        imported_count = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                print(f"Sheet {sheet_name} is empty. Skipping.")
                continue

            headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]
            while headers and headers[-1].startswith('Col') and headers[-1] != 'Col0':
                headers.pop()

            sheet_count = 0
            for r_idx, r in enumerate(rows[1:], start=2):
                if all(v is None for v in r):
                    continue
                doc = {
                    "category": sheet_name,
                    "created_at": datetime.datetime.utcnow()
                }
                for c_idx, val in enumerate(r):
                    if c_idx < len(headers):
                        header = headers[c_idx]
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            val = val.strftime("%Y-%m-%d")
                        doc[header] = val if val is not None else ""
                db.hosting_details.insert_one(doc)
                sheet_count += 1
                imported_count += 1
            print(f"Sheet '{sheet_name}': Imported {sheet_count} records.")

        print(f"\nMigration complete! Successfully imported {imported_count} total records to MongoDB Atlas.")

    except Exception as e:
        print(f"Failed to migrate updated hosting details: {e}")

if __name__ == '__main__':
    import_new_hosting()
