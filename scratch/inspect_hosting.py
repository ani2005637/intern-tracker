import openpyxl
import os

path = r"C:\Users\s.anirudh\Downloads\Domain & Hosting Account Details.xlsx"
if not os.path.exists(path):
    print(f"File not found at {path}")
    exit(1)

wb = openpyxl.load_workbook(path)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\nSheet: {name}, Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")
    # Print headers (row 1) and first row of data
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) > 0:
        print("Headers:", rows[0])
    if len(rows) > 1:
        print("Row 1 Data:", rows[1])
